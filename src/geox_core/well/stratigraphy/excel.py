"""
GEOX Well Stratigraphy — Generalized Excel + CSV Exporter
═══════════════════════════════════════════════════════════════════════════════

5-sheet XLSX output with conditional formatting.
Generalized from KL2 FINAL SOT.

DITEMPA BUKAN DIBERI — Forged, Not Given
"""

from __future__ import annotations

import csv
import logging
from collections import Counter

from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl

from .config import (
    ProjectConfig,
    MOTIF_COLORS,
    TRACT_COLORS,
    default_depo_eod,
)

logger = logging.getLogger("geox.stratigraphy.excel")

_THIN = Side(style="thin", color="BDBDBD")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Depo env colors for Excel
_DEPO_EOD = default_depo_eod()

# Well colors
_WELL_COLORS = {
    0: "#C5D9F1",
    1: "#E2EFDA",
    2: "#FCE4D6",
    3: "#FFF2CC",
    4: "#FFE0E0",
    5: "#EAD1DC",
    6: "#D9D2E9",
    7: "#CFE2F3",
}


def _thick_border():
    return Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header_row(ws, headers, row=1, color="0D47A1", font_color="FFFFFF"):
    """Write header row with styling."""
    fill = PatternFill("solid", fgColor=color)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = fill
        c.font = Font(color=font_color, bold=True, size=8)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _thick_border()
    ws.row_dimensions[row].height = 28


def _cell(ws, row, col, val, fill_color=None, bold=False):
    """Write a single cell."""
    c = ws.cell(row=row, column=col, value=val)
    if fill_color:
        fc = fill_color.lstrip("#")
        if len(fc) == 6:
            fc = "00" + fc  # add alpha
        try:
            c.fill = PatternFill("solid", fgColor=fc)
        except Exception:
            pass
    c.border = _thick_border()
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    if bold:
        c.font = Font(bold=True, size=8)
    return c


def export_excel(
    all_packages: list[dict],
    config: ProjectConfig,
    output_path: str,
) -> str:
    """
    Export 5-sheet XLSX: 01_GEO_PACKAGES, 02_10M_SENSING (placeholder),
    03_WELL_SUMMARY, 04_REGIONAL_SURFACES (placeholder), 05_COLOR_LEGEND.

    Returns output_path.
    """
    wb = openpyxl.Workbook()

    # ── 01_GEO_PACKAGES ──────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "01_GEO_PACKAGES"
    h1 = [
        "WELL",
        "PKG_ID",
        "PARENT_ZONE",
        "DEPO_ENV",
        "TOP_MD",
        "BASE_MD",
        "THICKNESS_M",
        "HUMAN_MOTIF",
        "RIDER_MOTIF",
        "NET_TREND",
        "VARIABILITY",
        "GR_BASELINE_SHIFT_GAPI",
        "GR_MEAN",
        "GR_P10",
        "GR_P50",
        "GR_P90",
        "LITHO",
        "SEQ_STRAT",
        "INFERRED_PROCESS",
        "ANOMALY_FLAG",
        "CONFIDENCE",
        "N_10M_BINS",
        "AGE_TOP_Ma",
        "AGE_BASE_Ma",
        "NN_ZONE",
        "QC_FLAGS",
        "EXTRAP_MOTIF",
        "EXTRAP_CONF",
    ]
    _header_row(ws1, h1, color="0D47A1")

    well_idx = {}
    for ri, p in enumerate(all_packages, 2):
        w = p["WELL"]
        if w not in well_idx:
            well_idx[w] = len(well_idx)
        wc = _WELL_COLORS.get(well_idx[w] % len(_WELL_COLORS), "FFFFFF")
        tf = _tract_fill(p.get("SEQ_STRAT", "UNKNOWN"))
        vals = [
            p["WELL"],
            p.get("PKG_ID", ""),
            p.get("PARENT_ZONE", ""),
            p.get("DEPO_ENV", ""),
            p["TOP"],
            p["BASE"],
            p["THICKNESS"],
            p["HUMAN_MOTIF"],
            p["RIDER_MOTIF"],
            p["NET_TREND"],
            p["VARIABILITY"],
            p["GR_BASELINE_SHIFT"],
            p["GR_MEAN"],
            p["GR_P10"],
            p["GR_P50"],
            p["GR_P90"],
            p.get("LITHO", ""),
            p.get("SEQ_STRAT", ""),
            p.get("INFERRED_PROCESS", ""),
            p.get("ANOMALY_FLAG", "PASS"),
            p.get("CONFIDENCE", "Medium"),
            p["N_BINS"],
            p.get("AGE_TOP_Ma"),
            p.get("AGE_BASE_Ma"),
            p.get("NN_ZONE", ""),
            p.get("QC", "PASS"),
            p.get("EXTRAP_MOTIF", ""),
            p.get("EXTRAP_CONF", ""),
        ]
        for ci, v in enumerate(vals, 1):
            fill = wc if ci == 1 else (tf if ci in (18, 19) else "FFFFFF")
            if ci in (27, 28) and p.get("EXTRAP_MOTIF"):
                fill = "FFF9C4"
            _cell(ws1, ri, ci, v, fill_color=fill)
        ws1.row_dimensions[ri].height = 18

    widths = [14, 12, 18, 14, 8, 8, 10, 18, 14, 18, 11, 14, 8, 8, 8, 8, 12, 8, 50, 20, 10, 8, 10, 10, 14, 10, 18, 10]
    for ci, w in enumerate(widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(h1))}1"

    # ── 03_WELL_SUMMARY ──────────────────────────────────────────────
    ws3 = wb.create_sheet("03_WELL_SUMMARY")
    h3 = [
        "WELL",
        "N_INTERVALS",
        "N_PACKAGES",
        "DOMINANT_MOTIF",
        "DOMINANT_TRACT",
        "DEPTH_RANGE_m",
        "AGE_RANGE_Ma",
        "DEPO_RANGE",
        "KEY_ANOMALIES",
        "GEOLOGICAL_INTERPRETATION",
    ]
    _header_row(ws3, h3, color="004D40")

    well_summaries = {}
    for p in all_packages:
        w = p["WELL"]
        if w not in well_summaries:
            well_summaries[w] = {
                "motifs": [],
                "tracts": [],
                "tops": [],
                "bases": [],
                "ages_top": [],
                "ages_base": [],
                "depos": set(),
                "flags": [],
            }
        s = well_summaries[w]
        s["motifs"].append(p["HUMAN_MOTIF"])
        s["tracts"].append(p.get("SEQ_STRAT", "UNKNOWN"))
        s["tops"].append(p["TOP"])
        s["bases"].append(p["BASE"])
        if p.get("AGE_TOP_Ma"):
            s["ages_top"].append(p["AGE_TOP_Ma"])
        if p.get("AGE_BASE_Ma"):
            s["ages_base"].append(p["AGE_BASE_Ma"])
        s["depos"].add(str(p.get("DEPO_ENV", "")).replace("2026_", ""))
        if p.get("ANOMALY_FLAG", "PASS") != "PASS":
            s["flags"].append(p["ANOMALY_FLAG"])

    well_order = config.well_order or list(dict.fromkeys(p["WELL"] for p in all_packages))
    for ri, well_id in enumerate(well_order, 2):
        s = well_summaries.get(well_id, {})
        ivs = config.intervals.get(well_id, [])
        wc = _WELL_COLORS.get((ri - 2) % len(_WELL_COLORS), "FFFFFF")
        dom_motif = Counter(s.get("motifs", [])).most_common(1)[0][0] if s.get("motifs") else "?"
        dom_tract = Counter(s.get("tracts", [])).most_common(1)[0][0] if s.get("tracts") else "?"
        d_rng = f"{min(s['tops']):.0f}-{max(s['bases']):.0f}m" if s.get("tops") and s.get("bases") else "?"
        a_rng = f"{min(s['ages_top']):.1f}-{max(s['ages_base']):.1f} Ma" if s.get("ages_top") and s.get("ages_base") else "?"
        d_envs = " -> ".join(sorted(s.get("depos", set())))
        flags = "; ".join(set(s.get("flags", []))) or "PASS"

        vals = [well_id, len(ivs), len(s.get("motifs", [])), dom_motif, dom_tract, d_rng, a_rng, d_envs, flags, ""]
        for ci, v in enumerate(vals, 1):
            _cell(ws3, ri, ci, v, fill_color=wc, bold=(ci == 1))
        ws3.row_dimensions[ri].height = 30

    widths3 = [14, 10, 10, 18, 10, 14, 14, 30, 22, 70]
    for ci, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # ── 05_COLOR_LEGEND ──────────────────────────────────────────────
    ws5 = wb.create_sheet("05_COLOR_LEGEND")
    _header_row(ws5, ["CATEGORY", "LABEL", "HEX_CODE", "DESCRIPTION", "RGB"], color="1A237E")

    legend_rows = []
    for m, c in MOTIF_COLORS.items():
        hc = c.lstrip("#")
        r, g, b = int(hc[0:2], 16), int(hc[2:4], 16), int(hc[4:6], 16)
        legend_rows.append(("MOTIF", m, c, f"GR motif: {m}", f"R{r} G{g} B{b}"))
    for t, c in TRACT_COLORS.items():
        hc = c.lstrip("#")
        r, g, b = int(hc[0:2], 16), int(hc[2:4], 16), int(hc[4:6], 16)
        legend_rows.append(("SEQ_STRAT", t, c, f"Systems tract: {t}", f"R{r} G{g} B{b}"))
    for code, (lbl, col) in _DEPO_EOD.items():
        hc = f"{int(col[0] * 255):02X}{int(col[1] * 255):02X}{int(col[2] * 255):02X}"
        legend_rows.append(("DEPO_ENV", code, f"#{hc}", lbl, f"R{int(col[0] * 255)} G{int(col[1] * 255)} B{int(col[2] * 255)}"))

    for ri, (cat, lbl, hx, desc, rgb) in enumerate(legend_rows, 2):
        hx_fill = hx.lstrip("#") if hx.startswith("#") and len(hx) == 7 else None
        for ci, v in enumerate([cat, lbl, hx, desc, rgb], 1):
            c = ws5.cell(row=ri, column=ci, value=v)
            if ci == 2 and hx_fill:
                try:
                    c.fill = PatternFill("solid", fgColor=hx_fill)
                except Exception:
                    pass
            c.border = _thick_border()
            c.alignment = Alignment(horizontal="left", vertical="center")
        ws5.row_dimensions[ri].height = 16

    for ci, w in enumerate([14, 24, 12, 50, 16], 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)
    logger.info(f"  XLSX: {output_path}")
    return output_path


def export_sensing(sensing_rows: list[dict], output_path: str) -> str:
    """Export 10 m sensing rows to CSV."""
    if not sensing_rows:
        with open(output_path, "w") as f:
            f.write("WELL,ZONE,TOP,BASE,N,P10,P50,P90,MEAN,RANGE,SLOPE,MICRO_MOTIF\n")
        return output_path

    fieldnames = ["WELL", "ZONE", "TOP", "BASE", "N", "P10", "P50", "P90", "MEAN", "RANGE", "SLOPE", "MICRO_MOTIF"]
    with open(output_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in sensing_rows:
            writer.writerow(row)

    logger.info(f"  CSV: {output_path}")
    return output_path


def _tract_fill(tract: str) -> str:
    """Return fill color for a systems tract."""
    return {
        "LST": "FFF3E0",
        "TST": "E3F2FD",
        "HST": "E8F5E9",
        "FSST": "FBE9E7",
        "CS": "F3E5F5",
        "UNCERTAIN": "ECEFF1",
        "UNKNOWN": "FAFAFA",
    }.get(tract, "FAFAFA")
